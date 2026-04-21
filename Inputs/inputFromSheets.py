import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

SHEET_NAME = "Formularsvar 1"
DEFINE_ANSWERS_SHEET = "Define Answers"
LIST_SPLIT_PATTERN = re.compile(r"[,;\n]+")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_list(value: Any) -> List[str]:
    if value is None:
        return []

    if isinstance(value, (list, tuple, set)):
        items = [_to_text(v) for v in value]
    else:
        text = _to_text(value)
        if not text:
            return []
        items = LIST_SPLIT_PATTERN.split(text)

    return [item.strip() for item in items if item and item.strip()]


def _normalize_kind(value: Any) -> str:
    text = _normalize_header(value).lower()
    if text.startswith("pref"):
        return "prefence"
    if text.startswith("trait"):
        return "traits"
    return text or "traits"


def _parse_weight(value: Any) -> int | float:
    text = _normalize_header(value)
    if not text:
        return 0

    try:
        numeric = float(text)
    except ValueError:
        return 0

    return int(numeric) if numeric.is_integer() else numeric


def _read_trait_catalog(
    workbook: Any,
    trait_columns: List[Dict[str, Any]],
    define_answers_sheet_name: str,
) -> List[Dict[str, Any]]:
    if define_answers_sheet_name not in workbook.sheetnames:
        return [
            {"index": index, "header": column_info["header"], "kind": "traits", "weight": 0}
            for index, column_info in enumerate(trait_columns)
        ]

    sheet = workbook[define_answers_sheet_name]
    trait_catalog: List[Dict[str, Any]] = []

    for index, column_info in enumerate(trait_columns):
        row_number = 2 + index
        trait_catalog.append(
            {
                "index": index,
                "header": column_info["header"],
                "kind": _normalize_kind(sheet.cell(row=row_number, column=2).value),
                "weight": _parse_weight(sheet.cell(row=row_number, column=3).value),
            }
        )

    return trait_catalog


def inputFromSheets(
    xlsx_path: str | Path,
    sheet_name: str = SHEET_NAME,
    define_answers_sheet_name: str = DEFINE_ANSWERS_SHEET,
) -> Dict[str, Any]:
    workbook = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' was not found in {xlsx_path}")

    sheet = workbook[sheet_name]

    trait_columns: List[Dict[str, Any]] = []
    for col in range(3, sheet.max_column + 1):
        header = _normalize_header(sheet.cell(row=1, column=col).value)
        if not header:
            continue
        trait_columns.append({"column": col, "header": header})

    trait_catalog = _read_trait_catalog(workbook, trait_columns, define_answers_sheet_name)

    people: List[Dict[str, Any]] = []
    for row in range(2, sheet.max_row + 1):
        person_id = _to_text(sheet.cell(row=row, column=2).value)
        if not person_id:
            break

        person: Dict[str, Any] = {"id": person_id, "attributes": []}

        for column_info in trait_columns:
            value = sheet.cell(row=row, column=column_info["column"]).value
            person["attributes"].append(_parse_list(value))

        people.append(person)

    result: Dict[str, Any] = {
        "schema_version": 1,
        "attribute_set": trait_catalog,
        "people": people,
    }

    xlsx_path = Path(xlsx_path)
    output_dir = Path(__file__).resolve().parent / "sheetOutput"
    people_output_path = output_dir / f"{xlsx_path.stem}wishes.json"
    trait_catalog_output_path = output_dir / f"{xlsx_path.stem}attribute_set.json"

    people_output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(people_output_path, "w", encoding="utf-8") as file:
        json.dump({"schema_version": 1, "people": people}, file, indent=4, ensure_ascii=False)

    trait_catalog_output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(trait_catalog_output_path, "w", encoding="utf-8") as file:
        json.dump({"schema_version": 1, "attribute_set": trait_catalog}, file, indent=4, ensure_ascii=False)

    return result


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert XLSX responses to project JSON input format")
    parser.add_argument("xlsx_path", help="Path to source .xlsx file")
    parser.add_argument(
        "--sheet",
        default=SHEET_NAME,
        help=f"Sheet name to parse (default: {SHEET_NAME})",
    )
    parser.add_argument(
        "--define-sheet",
        default=DEFINE_ANSWERS_SHEET,
        help=f"Sheet containing trait kinds in column B and weights in column C (default: {DEFINE_ANSWERS_SHEET})",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    result = inputFromSheets(
        args.xlsx_path,
        sheet_name=args.sheet,
        define_answers_sheet_name=args.define_sheet,
    )
    xlsx_path = Path(args.xlsx_path)
    output_dir = Path(__file__).resolve().parent / "sheetOutput"
    people_output = output_dir / f"{xlsx_path.stem}wishes.json"
    trait_catalog_output = output_dir / f"{xlsx_path.stem}attribute_set.json"
    print(f"Wrote {len(result['people'])} people to {people_output} and {trait_catalog_output}")


if __name__ == "__main__":
    main()
